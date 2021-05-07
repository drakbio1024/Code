import sys, glob, os
from winreg import *
import pandas as pd
import xlsxwriter
from openpyxl import *
from datetime import datetime

#script to get default downloads folder from windows registry
with OpenKey(HKEY_CURRENT_USER, 'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders') as key:
    Downloads = QueryValueEx(key, '{374DE290-123F-4565-9164-39C4925E467B}')[0]

all_downloads = glob.glob(Downloads + r'\*')

#Find latest full and open report from downloads folder,
print('    Reading reports from downloads.. may take a moment')

full_reports_list = []
for item in all_downloads:
    if item.startswith(Downloads + r'\Q2_Order_Report'):
        full_reports_list.append(item)
latest_full_report = max(full_reports_list, key = os.path.getctime)
full_df = pd.read_excel(latest_full_report, sheet_name='Report')

open_reports_list = []
for item in all_downloads:
    if item.startswith(Downloads + r'\Q2_Open_Order_Report'):
        open_reports_list.append(item)
latest_open_report = max(open_reports_list, key = os.path.getctime)
open_df = pd.read_excel(latest_open_report, sheet_name='Report')

#Set default path concocenated with backslash, for easier path naming
bs = "\\"
scriptpath = sys.path[0] + bs

#take input as Listname, add Listname to path to get full
#directory of desired spreadsheet to be updated.
print('Type filename of desired spreadsheet to import reports')
print('You can input either the name of a file (Print List 4-28) \
       \nor the full directory of the file. \nDo not include the file extention:.')

#check that the inputted file path exists, if it doesn't, ask again
test = True
while test == True:
    listname = input() + '.xlsx'
    if ':' in listname:
        listpath = listname
    else:
        listpath = (scriptpath + listname)
    if os.path.exists(listpath) == True:
        test = False
        print('Updating ' + listpath + '...')
    else:
        print('File not found. Please specify requested file \
               \nname inside the source directory of Update Print List.py:')

#export Full_df and Open_df to sheets in specified excel file.
while True:
    try:
        list_wb = load_workbook(listpath)
        break
    except PermissionError:
        print("Please close requested workbook, then press enter to continue:")
        input()

AddFull = True
AddOpen = True

if 'Open report' in list_wb.sheetnames:
    old_open_sheet = list_wb['Open report']
    old_open_in = True
    list_wb.remove(list_wb['Open report']) 
else:
    print('Open report not found in document. Would you like to add it? (y/n):')
    old_open_in = False
    while True:
        a = input().lower()
        if a == "y" or a == "yes":
            AddOpen = True
            break
        elif a == "n" or a == "no":
            AddOpen = False
            break
        else:
            print('Invalid input')
            print('Add open report? (y/n):')

if 'Full report' in list_wb.sheetnames:
    old_full_sheet = list_wb['Full report']
    old_full_in = True
    list_wb.remove(list_wb['Full report'])
else:
    print('Full report not found in document. Would you like to add it? (y/n):')
    old_full_in = False
    while True:
        a = input().lower()
        if a == "y" or a == "yes":
            AddFull = True
            break
        elif a == "n" or a == "no":
            AddFull = False
            break
        else:
            print('Invalid input')
            print('Add Full report? (y/n):')

options = {}
options['strings_to_formulas'] = False
options['strings_to_urls'] = False
while True:
    try:
        writer = pd.ExcelWriter(listpath, engine = 'openpyxl', mode = 'a', options = options)
        break
    except PermissionError:
        print('Please close requested workbook. Press enter to continue...')
        input()

print('Would you like to update the list of orders to print as well? (y/n) \nanswer no if object is not a print list')
yn = input().lower()
if yn == 'yes' or yn == 'y':
    print('Input date to pull TSD after: (MM/DD/YYYY) ')
    datei = input()
    while True:
        try:
            date = datetime.strptime(datei, '%m/%d/%Y')
            break
        except:
            print('Incorrect date format. enter as MM/DD/YYYY (Include 0s for M and D)')
            datei = input()
    print_df = open_df[open_df['Target Ship Date'] >= date]
    print_df = print_df[print_df['Internal Notes'] != 'Pending approved CI']
    print_df = print_df[(print_df['Order Status'] == 'In Process') & (print_df['Order Status'] == 'Ready')]
    print_df = print_df[pd.isnull(print_df['Last Order Pick Slip Print Date']) == True]
    print_df = print_df.reset_index()
    list_df = print_df[['PeopleSoft Order #']]
        
    while True:
        try:
            sheet = list_wb['Print list']
            break
        except:
            print('Print list sheet not found. Enter name of sheet to update or restart script and fix print list.')
            print('Here are sheets in the workbook: \n' + list_wb.sheetnames)
            sheet = list_wb[input()]
            break
    qrow = 2
    for i, r in list_df.iterrows():
        f = r.values
        sheet['A' + str(qrow)].value = f[0]
        qrow = qrow + 1

writer.book = list_wb
if AddFull == True:
    full_df.to_excel(writer, sheet_name = 'Full report', index = False)
    if old_full_in == True:
        list_wb['Full report'].column_dimensions = old_full_sheet.column_dimensions
    else:
        print('Column widths will be default for Full Report')
if AddOpen == True:
    open_df.to_excel(writer, sheet_name = 'Open report', index = False)
    if old_open_in == True:
        list_wb['Open report'].column_dimensions = old_open_sheet.column_dimensions
    else:
        print('Column widths will be default for Open Report')
list_wb.book = writer.book
list_wb.save(listpath)

print('Workbook saved. Press enter to continue...')
input()